#! python3
# multiplicationTable.py

import openpyxl, sys
from openpyxl.styles import Font
num = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
boldFont = Font(bold=True)
for i in range(1,num+1):
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i+1, column=1).font = boldFont
    sheet.cell(row=1, column=i+1).value = i
    sheet.cell(row=1, column=i+1).font = boldFont
for i in range(1,num+1):
    for j in range(1,num+1):
        sheet.cell(row=i+1, column=j+1).value = i*j
wb.save('multiplicationTable.xlsx')
