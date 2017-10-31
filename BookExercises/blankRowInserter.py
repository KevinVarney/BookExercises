#! python3
# blanRowInserter.py

import openpyxl, sys, os
if len(sys.argv) != 4:
    print('Incorrect number of arguments')
    exit(0)

# Get command line parameters
rowNum = int(sys.argv[1])
numRows = int(sys.argv[2])
wbName = sys.argv[3]

# Open source workbooks 
wb = openpyxl.load_workbook(wbName)
sheet = wb.get_active_sheet()

# Open destination workbook
wb2 = openpyxl.Workbook()
sheet2 = wb2.get_active_sheet()

# Copy across the first rows
for Row in sheet.iter_rows(min_row=0,max_row=rowNum):
    row_data = []
    for cell in Row:
        row_data.append(cell.value)
    sheet2.append(row_data)

# Insert some empty rows
for i in range(numRows):
    sheet2.append([])

# Copy across the rest of the rows
for Row in sheet.iter_rows(min_row=rowNum+1,max_row=sheet.max_row):
    row_data = []
    for cell in Row:
        row_data.append(cell.value)
    sheet2.append(row_data)

# Save the new workbook
wbNameParts = os.path.splitext(wbName)
wbName2 = wbNameParts[0] + 'Copy' + wbNameParts[1]
wb2.save(wbName2)

    
